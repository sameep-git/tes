from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import List, Optional
from collections import defaultdict
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..database import get_db
from .. import models, schemas

router = APIRouter(prefix="/api/schedules", tags=["schedules"])


def compute_section_numbers(sections) -> dict:
    """Compute section numbers for all sections in a schedule.

    Each timeslot has a base section_number (e.g. "010"). When multiple courses
    share the same timeslot, they get incrementing numbers from that base:
    010, 011, 012, etc.  Sections without a timeslot get None.

    Returns a dict mapping section.id -> computed section number string.
    """
    # Group sections by timeslot_id
    by_timeslot = defaultdict(list)
    for s in sections:
        if s.timeslot_id is not None:
            by_timeslot[s.timeslot_id].append(s)

    result = {}
    for _ts_id, group in by_timeslot.items():
        # Sort alphabetically by course code, then course name for determinism
        group.sort(key=lambda s: (
            s.course.code if s.course else "",
            s.course.name if s.course else "",
        ))

        # Get the base section number from the timeslot
        base = group[0].timeslot.section_number if group[0].timeslot else "000"
        base_int = int(base)

        for offset, sec in enumerate(group):
            result[sec.id] = f"{base_int + offset:03d}"

    return result


@router.get("/", response_model=List[schemas.ScheduleResponse])
def get_schedules(
    semester: Optional[str] = None,
    year: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
):
    q = db.query(models.Schedule).options(
        joinedload(models.Schedule.sections)
        .joinedload(models.Section.course),
        joinedload(models.Schedule.sections)
        .joinedload(models.Section.professor),
        joinedload(models.Schedule.sections)
        .joinedload(models.Section.timeslot),
        joinedload(models.Schedule.sections)
        .joinedload(models.Section.room),
    )
    if semester:
        q = q.filter(func.lower(models.Schedule.semester) == semester.lower())
    if year is not None:
        q = q.filter(models.Schedule.year == year)
    schedules = q.offset(skip).limit(limit).all()

    results = []
    for schedule in schedules:
        sec_nums = compute_section_numbers(schedule.sections)
        results.append(schemas.ScheduleResponse(
            id=schedule.id,
            semester=schedule.semester,
            year=schedule.year,
            status=schedule.status,
            finalized_at=schedule.finalized_at,
            sections=[
                schemas.SectionResponse.from_orm_with_relations(s, sec_nums.get(s.id))
                for s in schedule.sections
            ],
        ))
    return results


@router.get("/{schedule_id}/export")
def export_schedule_excel(
    schedule_id: int,
    db: Session = Depends(get_db),
):
    """Export a single schedule as an Excel (.xlsx) file."""
    schedule = db.query(models.Schedule).options(
        joinedload(models.Schedule.sections)
        .joinedload(models.Section.course),
        joinedload(models.Schedule.sections)
        .joinedload(models.Section.professor),
        joinedload(models.Schedule.sections)
        .joinedload(models.Section.timeslot),
        joinedload(models.Schedule.sections)
        .joinedload(models.Section.room),
    ).filter(models.Schedule.id == schedule_id).first()

    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")

    sec_nums = compute_section_numbers(schedule.sections)

    # Build section data sorted alphabetically by course code, then name
    rows = []
    for s in schedule.sections:
        rows.append({
            "section_number": sec_nums.get(s.id, ""),
            "course_code": s.course.code if s.course else "",
            "course_name": s.course.name if s.course else "",
            "professor": s.professor.name if s.professor else "Unassigned",
            "days": s.timeslot.days if s.timeslot else "",
            "start_time": s.timeslot.start_time if s.timeslot else "",
            "end_time": s.timeslot.end_time if s.timeslot else "",
            "room": f"{s.room.building} {s.room.room_number}" if s.room else "",
            "status": s.status,
        })

    rows.sort(key=lambda r: (r["course_code"], r["course_name"]))

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{schedule.semester} {schedule.year}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    headers = [
        "Section", "Course Code", "Course Name", "Professor",
        "Days", "Start Time", "End Time", "Room", "Status"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        values = [
            row_data["section_number"],
            row_data["course_code"],
            row_data["course_name"],
            row_data["professor"],
            row_data["days"],
            row_data["start_time"],
            row_data["end_time"],
            row_data["room"],
            row_data["status"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 1:  # Section number column
                cell.alignment = Alignment(horizontal="center")

    # Auto-size columns
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{schedule.semester}_{schedule.year}_Schedule.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
